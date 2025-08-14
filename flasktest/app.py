from flask import Flask, render_template, request
from flask import send_file
from xhtml2pdf import pisa
from flask import send_file, request
from flask import Response, render_template_string
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import io
from utils.api import (
    get_token,
    get_project_list,
    get_project_list,
    get_fpy,
    get_station_ntf_details,
    get_station_der_details
)
from utils.helpers import get_top_n_counts
from config import NTF_GOALS, DER_GOALS
from datetime import datetime
import pandas as pd

app = Flask(__name__)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/auto-data')
def auto_data():
    token = get_token()
    projects = get_project_list(token)
    fpy_data = get_fpy(token, projects)

    desired_columns = ["project", "station", "inPut", "pass", "fail", "notFail", "der", "ntf", "rty"]
    filtered_data = [
        {col: row.get(col, "") for col in desired_columns}
        for row in fpy_data
    ]

    current_time = datetime.now().strftime("%H:%M")
    return render_template('auto_data.html', data=filtered_data, current_time=current_time)

@app.route('/project-specific', methods=['GET', 'POST'])
def project_specific():
    token = get_token()
    projects = get_project_list(token)
    selected_project = None
    rty_goal = 90.0
    fpy_data = []
    failed_stations = []
    fail_details = []

    if request.method == 'POST':
        selected_project = request.form.get('project')
        rty_goal = float(request.form.get('rty_goal', 90.0))
        fpy_data_raw = get_fpy(token, [selected_project])

        desired_columns = ["project", "station", "inPut", "pass", "fail", "notFail", "der", "ntf", "rty"]
        fpy_data = [
            {col: row.get(col, "") for col in desired_columns}
            for row in fpy_data_raw
        ]

        if fpy_data and "rty" in fpy_data[0]:
            try:
                actual_rty = float(str(fpy_data[0]["rty"]).replace("%", ""))
                if actual_rty < rty_goal:
                    for row in fpy_data:
                        station = row.get("station")
                        ntf = float(str(row.get("ntf", "0")).replace("%", "")) if row.get("ntf") else None
                        der = float(str(row.get("der", "0")).replace("%", "")) if row.get("der") else None

                        if station in NTF_GOALS and ntf is not None and ntf > NTF_GOALS[station]:
                            failed_stations.append((station, "NTF", ntf, NTF_GOALS[station]))
                            detail_data = get_station_ntf_details(token, selected_project, station)
                            detail_df = pd.DataFrame(detail_data)
                            detail_df = detail_df.rename(columns={
                                "substation": "Computer Name",
                                "sn": "SN",
                                "symptomEnName": "Fault Description"
                            })
                            detail_df = detail_df[["SN", "Fault Description", "Computer Name"]]

                            top_computers = detail_df["Computer Name"].value_counts().head(3).to_dict()
                            top_faults_by_computer = {}
                            for comp in top_computers:
                                comp_faults = detail_df[detail_df["Computer Name"] == comp]
                                faults = comp_faults["Fault Description"].value_counts().head(3).reset_index().values.tolist()
                                top_faults_by_computer[comp] = faults

                            fail_details.append({
                                "station": station,
                                "metric": "NTF",
                                "actual": ntf,
                                "goal": NTF_GOALS[station],
                                "top_computers": top_computers,
                                "top_faults_by_computer": top_faults_by_computer
                            })

                        if station in DER_GOALS and der is not None and der > DER_GOALS[station]:
                            failed_stations.append((station, "DER", der, DER_GOALS[station]))
                            detail_data = get_station_der_details(token, selected_project, station)
                            detail_df = pd.DataFrame(detail_data)
                            detail_df = detail_df.rename(columns={
                                "sn": "SN",
                                "responsibilityEnName": "Responsibility",
                                "symptomEnName": "Symptoms"
                            })
                            detail_df = detail_df[["SN", "Responsibility", "Symptoms"]]
                            top_symptoms = get_top_n_counts(detail_df, "Symptoms", 3)
                            top_responsibilities = get_top_n_counts(detail_df, "Responsibility", 3)

                            fail_details.append({
                                "station": station,
                                "metric": "DER",
                                "actual": der,
                                "goal": DER_GOALS[station],
                                "top_symptoms": top_symptoms.to_dict(orient="records"),
                                "top_responsibilities": top_responsibilities.to_dict(orient="records")
                            })
            except Exception as e:
                print("RTY analysis error:", e)

    return render_template("project_specific.html",
                           projects=projects,
                           selected_project=selected_project,
                           rty_goal=rty_goal,
                           data=fpy_data,
                           failed_stations=failed_stations,
                           fail_details=fail_details)



@app.route('/multi-project-goals', methods=['GET', 'POST'])
def multi_project_goals():
    token = get_token()
    projects = get_project_list(token)
    goals = {p: 90.0 for p in projects}
    results = {}

    if request.method == 'POST':
        goals = {p: float(request.form.get(f'goal_{p}', 90.0)) for p in projects}
        for project in projects:
            fpy_data = get_fpy(token, [project])
            if fpy_data and "rty" in fpy_data[0]:
                actual_rty = float(str(fpy_data[0]["rty"]).replace("%", ""))
                status = "Pass" if actual_rty >= goals[project] else "Fail"
                results[project] = f"{status} — Actual: {actual_rty:.2f}% | Goal: {goals[project]:.2f}%"
            else:
                results[project] = "No RTY data"

    return render_template('multi_project_goals.html', goals=goals, results=results)

@app.route('/top-failures')
def top_failures():
    token = get_token()
    projects = get_project_list(token)
    return render_template('top_failures.html', projects=projects)

from flask import send_file, request
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

@app.route('/export-excel')
def export_excel():
    project = request.args.get('project')
    rty_goal = float(request.args.get('rty_goal', 90.0))

    token = get_token()
    fpy_data_raw = get_fpy(token, [project])

    if not fpy_data_raw:
        return "No data to export."

    # Clean FPY table
    desired_columns = ["project", "station", "inPut", "pass", "fail", "notFail", "der", "ntf", "rty"]
    fpy_data = [{col: row.get(col, "") for col in desired_columns} for row in fpy_data_raw]
    fpy_df = pd.DataFrame(fpy_data).astype(str)

    failed_stations = []
    ntf_rows = []
    der_rows = []

    try:
        actual_rty = float(str(fpy_df["rty"].iloc[0]).replace("%", ""))
        if actual_rty < rty_goal:
            for _, row in fpy_df.iterrows():
                station = row["station"]
                ntf = float(str(row["ntf"]).replace("%", "")) if row["ntf"] else None
                der = float(str(row["der"]).replace("%", "")) if row["der"] else None

                if station in NTF_GOALS and ntf > NTF_GOALS[station]:
                    failed_stations.append((station, "NTF", ntf, NTF_GOALS[station]))
                    detail_df = pd.DataFrame(get_station_ntf_details(token, project, station)).rename(columns={
                        "substation": "Computer Name",
                        "sn": "SN",
                        "symptomEnName": "Fault Description"
                    })[["SN", "Fault Description", "Computer Name"]]

                    top_computers = detail_df["Computer Name"].value_counts().head(3).to_dict()
                    for comp, count in top_computers.items():
                        faults = detail_df[detail_df["Computer Name"] == comp]["Fault Description"].value_counts().head(3)
                        fault_lines = [f"{i+1}. {fault} → {qty}" for i, (fault, qty) in enumerate(faults.items())]
                        ntf_rows.append([f"{comp} → {count}", "\n".join(fault_lines)])

                if station in DER_GOALS and der > DER_GOALS[station]:
                    failed_stations.append((station, "DER", der, DER_GOALS[station]))
                    detail_df = pd.DataFrame(get_station_der_details(token, project, station)).rename(columns={
                        "sn": "SN",
                        "responsibilityEnName": "Responsibility",
                        "symptomEnName": "Symptoms"
                    })[["SN", "Responsibility", "Symptoms"]]

                    top_symptoms = detail_df["Symptoms"].value_counts().head(3)
                    top_responsibilities = detail_df["Responsibility"].value_counts().head(3)

                    for i in range(3):
                        symptom = top_symptoms.index[i] if i < len(top_symptoms) else ""
                        symptom_qty = top_symptoms.iloc[i] if i < len(top_symptoms) else ""
                        resp = top_responsibilities.index[i] if i < len(top_responsibilities) else ""
                        resp_qty = top_responsibilities.iloc[i] if i < len(top_responsibilities) else ""
                        der_rows.append([f"{symptom} → {symptom_qty}", f"{resp} → {resp_qty}"])
    except Exception as e:
        print("RTY analysis error:", e)

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "FPY Report"

    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def write_section(title, df_or_rows, headers=None):
        # Safely merge title row
        pre_row = ws.max_row if ws.max_row else 0
        ws.append([title])
        title_row = pre_row + 1
        ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=10)
        ws.cell(row=title_row, column=1).font = bold

        if isinstance(df_or_rows, pd.DataFrame):
            ws.append(list(df_or_rows.columns))
            for cell in ws[ws.max_row]:
                cell.font = bold
                cell.alignment = center
            for row in df_or_rows.itertuples(index=False):
                ws.append(list(row))
        else:
            if headers:
                ws.append(headers)
                for cell in ws[ws.max_row]:
                    cell.font = bold
                    cell.alignment = center
            for row in df_or_rows:
                ws.append(row)

        ws.append([])  # Spacer row

    # Write FPY Table
    write_section("FPY Table", fpy_df)

    # Write Failed Stations
    if failed_stations:
        fail_df = pd.DataFrame(failed_stations, columns=["Station", "Metric", "Actual (%)", "Goal (%)"])
        write_section("Failed Stations", fail_df)

    # Write NTF Breakdown
    if ntf_rows:
        write_section("Top Failure Analysis — NTF", ntf_rows, ["Top Computer → Qty", "Top 3 Faults → Qty"])

    # Write DER Breakdown
    if der_rows:
        write_section("Top Failure Analysis — DER", der_rows, ["Symptom → Qty", "Responsibility → Qty"])

    # Adjust column width for 'project' column
    ws.column_dimensions['A'].width = 25  # Wider than others

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output,
                     download_name=f"{project}_full_report.xlsx",
                     as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')



@app.route('/export-pdf')
def export_pdf():
    project = request.args.get('project')
    rty_goal = float(request.args.get('rty_goal', 90.0))

    token = get_token()
    fpy_data_raw = get_fpy(token, [project])

    if not fpy_data_raw:
        return "No data to export."

    # Clean FPY table
    desired_columns = ["project", "station", "inPut", "pass", "fail", "notFail", "der", "ntf", "rty"]
    fpy_data = [{col: row.get(col, "") for col in desired_columns} for row in fpy_data_raw]
    fpy_df = pd.DataFrame(fpy_data).astype(str)

    failed_stations = []
    ntf_blocks = []
    der_blocks = []

    try:
        actual_rty = float(str(fpy_df["rty"].iloc[0]).replace("%", ""))
        if actual_rty < rty_goal:
            for _, row in fpy_df.iterrows():
                station = row.get("station")
                ntf = float(str(row.get("ntf", "0")).replace("%", "")) if row.get("ntf") else None
                der = float(str(row.get("der", "0")).replace("%", "")) if row.get("der") else None

                if station in NTF_GOALS and ntf is not None and ntf > NTF_GOALS[station]:
                    failed_stations.append((station, "NTF", ntf, NTF_GOALS[station]))
                    detail_data = get_station_ntf_details(token, project, station)
                    detail_df = pd.DataFrame(detail_data).rename(columns={
                        "substation": "Computer Name",
                        "sn": "SN",
                        "symptomEnName": "Fault Description"
                    })
                    detail_df = detail_df[["SN", "Fault Description", "Computer Name"]]

                    top_computers = detail_df["Computer Name"].value_counts().head(3).to_dict()
                    rows = ""
                    for comp, count in top_computers.items():
                        comp_faults = detail_df[detail_df["Computer Name"] == comp]
                        faults = comp_faults["Fault Description"].value_counts().head(3)
                        fault_lines = "".join([f"{i+1}. {fault} → {qty}<br>" for i, (fault, qty) in enumerate(faults.items())])
                        rows += f"<tr><td>{comp} → {count}</td><td>{fault_lines}</td></tr>"
                    ntf_blocks.append(f"""
                        <h3>{station} — NTF Analysis</h3>
                        <table>
                            <thead><tr><th>Top Computer → Qty</th><th>Top 3 Faults → Qty</th></tr></thead>
                            <tbody>{rows}</tbody>
                        </table>
                    """)

                if station in DER_GOALS and der is not None and der > DER_GOALS[station]:
                    failed_stations.append((station, "DER", der, DER_GOALS[station]))
                    detail_df = pd.DataFrame(get_station_der_details(token, project, station)).rename(columns={
                        "sn": "SN",
                        "responsibilityEnName": "Responsibility",
                        "symptomEnName": "Symptoms"
                    })
                    detail_df = detail_df[["SN", "Responsibility", "Symptoms"]]

                    top_symptoms = detail_df["Symptoms"].value_counts().head(3)
                    top_responsibilities = detail_df["Responsibility"].value_counts().head(3)

                    rows = ""
                    for i in range(3):
                        symptom = top_symptoms.index[i] if i < len(top_symptoms) else ""
                        symptom_qty = top_symptoms.iloc[i] if i < len(top_symptoms) else ""
                        resp = top_responsibilities.index[i] if i < len(top_responsibilities) else ""
                        resp_qty = top_responsibilities.iloc[i] if i < len(top_responsibilities) else ""
                        rows += f"<tr><td>{symptom} → {symptom_qty}</td><td>{resp} → {resp_qty}</td></tr>"
                    der_blocks.append(f"""
                        <h3>{station} — DER Analysis</h3>
                        <table>
                            <thead><tr><th>Symptom → Qty</th><th>Responsibility → Qty</th></tr></thead>
                            <tbody>{rows}</tbody>
                        </table>
                    """)
    except Exception as e:
        print("RTY analysis error:", e)

    # Build FPY table manually with wider project column
    fpy_html = "<table border='1' cellspacing='0' cellpadding='4' style='width:100%;'>"
    fpy_html += "<thead><tr>"
    for col in fpy_df.columns:
        if col == "project":
            fpy_html += f"<th style='font-size:10px; white-space:normal; word-wrap:break-word; width:120px;'>{col}</th>"
        else:
            fpy_html += f"<th style='font-size:10px; white-space:normal; word-wrap:break-word;'>{col}</th>"
    fpy_html += "</tr></thead><tbody>"

    for _, row in fpy_df.iterrows():
        fpy_html += "<tr>"
        for col in fpy_df.columns:
            if col == "project":
                fpy_html += f"<td style='font-size:10px; white-space:normal; word-wrap:break-word; width:120px;'>{row[col]}</td>"
            else:
                fpy_html += f"<td style='font-size:10px; white-space:normal; word-wrap:break-word;'>{row[col]}</td>"
        fpy_html += "</tr>"
    fpy_html += "</tbody></table>"

    # Build full HTML
    html = f"""
    <html><head><style>
    body {{ font-family: Arial; font-size: 11px; }}
    table {{ border-collapse: collapse; width: 100%; margin-bottom: 20px; }}
    th, td {{ border: 1px solid #ccc; padding: 4px; text-align: left; vertical-align: top; }}
    th {{ background-color: #f0f0f0; }}
    h1 {{ margin-bottom: 10px; }}
    h2 {{ margin-top: 30px; }}
    h3 {{ margin-top: 20px; }}
    </style></head><body>
    <h1>FPY Report for {project}</h1>
    <p><strong>RTY Goal:</strong> {rty_goal}%</p>

    <h2>FPY Table</h2>
    {fpy_html}
    """

    if failed_stations:
        fail_df = pd.DataFrame(failed_stations, columns=["Station", "Metric", "Actual (%)", "Goal (%)"])
        html += "<h2>Failed Stations</h2>" + fail_df.to_html(index=False)

    if ntf_blocks or der_blocks:
        html += "<h2>Top Failure Analysis</h2>"
        html += "".join(ntf_blocks)
        html += "".join(der_blocks)

    html += "</body></html>"

    # Convert to PDF
    pdf = io.BytesIO()
    pisa.CreatePDF(io.StringIO(html), dest=pdf)
    pdf.seek(0)
    return send_file(pdf,
                     download_name=f"{project}_full_report.pdf",
                     as_attachment=True,
                     mimetype='application/pdf')


@app.route('/project-range-report', methods=['POST'])
def project_range_report():
    data = request.get_json()
    project = data.get("project")
    start_date = data.get("start_date")
    end_date = data.get("end_date")
    rty_goal = float(data.get("rty_goal", 90.0))

    token = get_token()
    fpy_data_raw = get_fpy(token, [project], start_date=start_date, end_date=end_date)

    if not fpy_data_raw:
        return {"error": "No data found for given range"}, 404

    desired_columns = ["project", "station", "inPut", "pass", "fail", "notFail", "der", "ntf", "rty"]
    fpy_data = [{col: row.get(col, "") for col in desired_columns} for row in fpy_data_raw]
    fpy_df = pd.DataFrame(fpy_data).astype(str)

    failed_stations = []
    ntf_rows = []
    der_rows = []

    try:
        actual_rty = float(str(fpy_df["rty"].iloc[0]).replace("%", ""))
        if actual_rty < rty_goal:
            for _, row in fpy_df.iterrows():
                station = row["station"]
                ntf = float(str(row["ntf"]).replace("%", "")) if row["ntf"] else None
                der = float(str(row["der"]).replace("%", "")) if row["der"] else None

                if station in NTF_GOALS and ntf > NTF_GOALS[station]:
                    failed_stations.append((station, "NTF", ntf, NTF_GOALS[station]))
                    detail_df = pd.DataFrame(get_station_ntf_details(token, project, station, start_date, end_date)).rename(columns={
                        "substation": "Computer Name",
                        "sn": "SN",
                        "symptomEnName": "Fault Description"
                    })[["SN", "Fault Description", "Computer Name"]]

                    top_computers = detail_df["Computer Name"].value_counts().head(3).to_dict()
                    for comp, count in top_computers.items():
                        faults = detail_df[detail_df["Computer Name"] == comp]["Fault Description"].value_counts().head(3)
                        fault_lines = [f"{i+1}. {fault} → {qty}" for i, (fault, qty) in enumerate(faults.items())]
                        ntf_rows.append([f"{comp} → {count}", fault_lines])

                if station in DER_GOALS and der > DER_GOALS[station]:
                    failed_stations.append((station, "DER", der, DER_GOALS[station]))
                    detail_df = pd.DataFrame(get_station_der_details(token, project, station, start_date, end_date)).rename(columns={
                        "sn": "SN",
                        "responsibilityEnName": "Responsibility",
                        "symptomEnName": "Symptoms"
                    })[["SN", "Responsibility", "Symptoms"]]

                    top_symptoms = detail_df["Symptoms"].value_counts().head(3)
                    top_responsibilities = detail_df["Responsibility"].value_counts().head(3)

                    for i in range(3):
                        symptom = top_symptoms.index[i] if i < len(top_symptoms) else ""
                        symptom_qty = top_symptoms.iloc[i] if i < len(top_symptoms) else ""
                        resp = top_responsibilities.index[i] if i < len(top_responsibilities) else ""
                        resp_qty = top_responsibilities.iloc[i] if i < len(top_responsibilities) else ""
                        der_rows.append([f"{symptom} → {symptom_qty}", f"{resp} → {resp_qty}"])
    except Exception as e:
        print("RTY analysis error:", e)

    return {
        "project": project,
        "start_date": start_date,
        "end_date": end_date,
        "rty_goal": rty_goal,
        "fpy_table": fpy_df.to_dict(orient="records"),
        "failed_stations": failed_stations,
        "ntf_analysis": ntf_rows,
        "der_analysis": der_rows
    }



if __name__ == '__main__':
    app.run(host="0.0.0.0", port=5000)


